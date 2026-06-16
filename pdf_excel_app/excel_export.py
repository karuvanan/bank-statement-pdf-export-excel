from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


TRANSACTION_HEADERS = [
    "#",
    "Source PDF",
    "Page",
    "Date",
    "Desc",
    "Details Payee Reference",
    "Amount",
    "Debit",
    "Credit",
    "Statement Balance",
    "OCR Amount",
    "OCR Balance",
    "Review Notes",
]

SUMMARY_HEADERS = [
    "Source PDF",
    "Period",
    "Beginning Balance",
    "Ending Balance",
    "Transactions",
    "Calculated Debit",
    "Statement Total Debit",
    "Calculated Credit",
    "Statement Total Credit",
    "Rows With Notes",
]


def split_description(value: str | None) -> tuple[str, str]:
    parts = [part.strip() for part in str(value or "").splitlines() if part.strip()]
    desc = parts[0] if parts else ""
    details = " ".join(parts[1:]) or desc
    return desc, details


def parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", str(value))
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def safe_sheet_name(name: str, existing: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", " ", re.sub(r"\.pdf$", "", name, flags=re.I)).strip() or "Sheet"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in existing:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def money_or_blank(value: Any) -> float | None:
    return None if value is None else float(value)


def transaction_row(row: dict[str, Any], index: int) -> list[Any]:
    desc, details = split_description(row.get("description"))
    return [
        index,
        row.get("source_pdf", ""),
        row.get("page"),
        parse_iso_date(row.get("entry_date_full")),
        desc,
        details,
        money_or_blank(row.get("amount")),
        money_or_blank(row.get("debit")),
        money_or_blank(row.get("credit")),
        money_or_blank(row.get("balance")),
        row.get("amount_text", ""),
        row.get("balance_text", ""),
        row.get("notes", ""),
    ]


def add_table(ws, name: str, max_row: int, max_col: int) -> None:
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table_name = re.sub(r"[^A-Za-z0-9_]", "", name)
    if not table_name or table_name[0].isdigit():
        table_name = f"Table{table_name}"
    table = Table(displayName=table_name[:240], ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws, max_row: int, max_col: int, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="174A7C")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def highlight_missing_amounts(ws, max_row: int) -> None:
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_index in range(2, max_row + 1):
        desc = ws.cell(row_index, 5).value
        details = ws.cell(row_index, 6).value
        amount = ws.cell(row_index, 7).value
        is_beginning_balance = str(desc or "").strip().upper() == "BEGINNING BALANCE"
        if not is_beginning_balance and desc not in (None, "") and details not in (None, "") and amount in (None, ""):
            for col_index in (7, 8, 9):
                ws.cell(row_index, col_index).fill = warning_fill


def add_transaction_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], existing: set[str]) -> None:
    ws = wb.create_sheet(safe_sheet_name(title, existing))
    ws.append(TRANSACTION_HEADERS)
    for index, row in enumerate(rows, start=1):
        ws.append(transaction_row(row, index))

    max_row = max(ws.max_row, 1)
    max_col = len(TRANSACTION_HEADERS)
    style_sheet(ws, max_row, max_col, [6, 16, 8, 13, 24, 48, 14, 14, 14, 16, 16, 16, 24])
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        row[3].number_format = "dd/mm/yyyy"
        for idx in (6, 7, 8, 9):
            row[idx].number_format = "#,##0.00"
    highlight_missing_amounts(ws, max_row)
    if max_row >= 2:
        add_table(ws, f"{ws.title}Table", max_row, max_col)


def add_summary_sheet(wb: Workbook, summaries: list[dict[str, Any]], existing: set[str]) -> None:
    ws = wb.active
    ws.title = safe_sheet_name("Summary", existing)
    ws.append(SUMMARY_HEADERS)
    for item in summaries:
        ws.append(
            [
                item.get("source_pdf"),
                item.get("statement_period"),
                money_or_blank(item.get("beginning_balance")),
                money_or_blank(item.get("ending_balance")),
                item.get("transaction_count"),
                money_or_blank(item.get("calculated_debit")),
                money_or_blank(item.get("statement_total_debit")),
                money_or_blank(item.get("calculated_credit")),
                money_or_blank(item.get("statement_total_credit")),
                item.get("rows_with_notes"),
            ]
        )
    max_row = max(ws.max_row, 1)
    max_col = len(SUMMARY_HEADERS)
    style_sheet(ws, max_row, max_col, [18, 12, 18, 18, 14, 18, 20, 18, 20, 16])
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for idx in (2, 3, 5, 6, 7, 8):
            row[idx].number_format = "#,##0.00"
    if max_row >= 2:
        add_table(ws, "StatementSummaryTable", max_row, max_col)


def export_workbook(payload: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(payload["rows"], key=lambda row: (row["source_pdf"], row["page"], row["y"]))
    summaries = payload["statement_summaries"]

    wb = Workbook()
    existing: set[str] = set()
    add_summary_sheet(wb, summaries, existing)
    add_transaction_sheet(wb, "All Transactions", rows, existing)

    by_pdf: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_pdf.setdefault(row["source_pdf"], []).append(row)
    for pdf_name, pdf_rows in by_pdf.items():
        add_transaction_sheet(wb, pdf_name, pdf_rows, existing)

    wb.save(output_path)
    return output_path
